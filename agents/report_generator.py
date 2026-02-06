import pandas as pd
from database.models import Analysis, get_session
from utils.helpers import logger, get_cairo_time
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def generate_excel_report():
    session = get_session()
    date_str = get_cairo_time().strftime('%Y-%m-%d')
    
    # جلب التحليلات
    analyses = session.query(Analysis).all()
    
    if not analyses:
        logger.warning("No analysis data found")
        return None
    
    # إنشاء Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Analysis Report"
    
    # العنوان
    headers = ['Symbol', 'Price', 'Signal', 'Score', 'Trend', 'RSI', 
               'SMA 20', 'SMA 50', 'Target 1', 'Target 2', 'Stop Loss']
    
    header_fill = PatternFill(start_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # البيانات
    for idx, a in enumerate(analyses, 2):
        ws.cell(row=idx, column=1, value=a.symbol)
        ws.cell(row=idx, column=2, value=float(a.current_price))
        ws.cell(row=idx, column=3, value=a.signal)
        ws.cell(row=idx, column=4, value=a.score)
        ws.cell(row=idx, column=5, value=a.trend)
        ws.cell(row=idx, column=6, value=float(a.rsi) if a.rsi else 0)
        ws.cell(row=idx, column=7, value=float(a.sma_20))
        ws.cell(row=idx, column=8, value=float(a.sma_50))
        ws.cell(row=idx, column=9, value=float(a.target_1))
        ws.cell(row=idx, column=10, value=float(a.target_2))
        ws.cell(row=idx, column=11, value=float(a.stop_loss))
        
        # تلوين الإشارة
        signal_cell = ws.cell(row=idx, column=3)
        if a.signal == 'BUY':
            signal_cell.fill = PatternFill(start_color="90EE90", fill_type="solid")
        elif a.signal == 'SELL':
            signal_cell.fill = PatternFill(start_color="FF6B6B", fill_type="solid")
    
    # تعديل العرض
    for col in range(1, 12):
        ws.column_dimensions[chr(64+col)].width = 15
    
    # حفظ
    output_dir = "/tmp/reports"
    os.makedirs(output_dir, exist_ok=True)
    filepath = f"{output_dir}/Report_{date_str}.xlsx"
    wb.save(filepath)
    
    logger.info(f"✅ Excel report saved: {filepath}")
    return filepath